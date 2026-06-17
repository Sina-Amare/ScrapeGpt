"""Project workflow endpoints."""

import csv
import io
import json
import logging
import time
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    HTTPException,
    Query,
    Request,
    Response,
    status,
)
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db
from app.core.rate_limit import SCRAPE_RATE_LIMIT, limiter
from app.models.job import (
    ACTIVE_EXTRACTION_RUN_STATES,
    ACTIVE_PROJECT_STATES,
    DELETABLE_PROJECT_STATES,
    CrawlPage,
    CrawlPageState,
    Export,
    ExtractedRecord,
    ExtractionMode,
    ExtractionRun,
    ExtractionSpec,
    Project,
    ProjectState,
)
from app.models.user import User
from app.schemas.project import (
    BlockedPageDetail,
    ExtractRequest,
    ExtractionProgress,
    ExtractionQuality,
    ExtractionSpecResponse,
    ExtractionSpecUpdate,
    FrontierPreviewResponse,
    FrontierUrlDecision,
    PreviewResponse,
    ProjectAnalyzeRequest,
    ProjectEventResponse,
    ProjectListItem,
    ProjectResponse,
    RecordPageResponse,
    RecordResponse,
    RetryRequest,
)
from app.services.crawl_scope import ScopeConfirmationError
from app.services.extraction_mode import resolve_extraction_mode
from app.services.fetcher import FetchError, fetch_url
from app.services.interaction_detect import detect_interaction_profile
from app.services.interaction_profile import is_enabled as _interactions_enabled
from app.services.interaction_profile import merge_enabled as _interactions_merge_enabled
from app.services.interaction_profile import metadata_columns as _interaction_metadata_columns
from app.services.session_service import get_cookies_for_session
from app.services.url_validator import URLValidationError, validate_url
from app.services.project_events import list_project_events, record_project_event
from app.services.extraction_spec_service import (
    ensure_default_spec,
    flag_invalid_field_selectors,
    latest_spec,
    selected_field_count,
)
from app.services.frontierpreview import create_frontier_preview, latest_frontier_preview
from app.services.job_admission import JobAdmissionError, JobAdmissionErrorType, admit_job
from app.services.job_executor import execute_job_pipeline
from app.services.job_state import transition_job_to_canceled
from app.services.project_lifecycle import delete_project_tree
from app.services.project_extraction import (
    ExtractionAlreadyRunningError,
    count_records,
    execute_project_extraction,
    list_records,
    start_project_extraction,
)
from app.services.project_preview import create_preview, latest_preview
from app.services.project_retry import RetryError, retry_failed_project
from app.services.project_status import confidence_label, detected_type, product_status_for

router = APIRouter(prefix="/projects", tags=["Projects"])

logger = logging.getLogger(__name__)


async def _display_run_id(db: AsyncSession, project_id: int) -> int | None:
    """The run whose progress/pages we surface: the active run if one is in
    flight (so live progress shows), else the project's current completed run."""
    active = await db.scalar(
        select(ExtractionRun.id).where(
            ExtractionRun.project_id == project_id,
            ExtractionRun.state.in_(ACTIVE_EXTRACTION_RUN_STATES),
        )
    )
    if active is not None:
        return active
    return await db.scalar(
        select(Project.current_extraction_run_id).where(Project.id == project_id)
    )


async def _progress(db: AsyncSession, project_id: int) -> ExtractionProgress:
    run_id = await _display_run_id(db, project_id)
    if run_id is None:
        return ExtractionProgress()
    crawl_pages_total = await db.scalar(
        select(func.count(CrawlPage.id)).where(
            CrawlPage.extraction_run_id == run_id
        )
    )
    page_counts_result = await db.execute(
        select(CrawlPage.state, func.count(CrawlPage.id))
        .where(CrawlPage.extraction_run_id == run_id)
        .group_by(CrawlPage.state)
    )
    page_counts = {
        state.value if hasattr(state, "value") else str(state): int(count)
        for state, count in page_counts_result
    }
    records = await db.scalar(
        select(func.count(ExtractedRecord.id)).where(
            ExtractedRecord.extraction_run_id == run_id
        )
    )
    exports = await db.scalar(
        select(func.count(Export.id)).where(Export.extraction_run_id == run_id)
    )

    blocked_rows = (await db.execute(
        select(CrawlPage.normalized_url, CrawlPage.block_reason, CrawlPage.error)
        .where(
            CrawlPage.extraction_run_id == run_id,
            CrawlPage.state == CrawlPageState.BLOCKED,
        )
        .limit(100)
    )).all()
    blocked_detail = [
        BlockedPageDetail(
            url=row.normalized_url or "",
            block_reason=row.block_reason or "UNKNOWN",
            error=row.error,
        )
        for row in blocked_rows
    ]

    failed_rows = (await db.execute(
        select(CrawlPage.normalized_url, CrawlPage.block_reason, CrawlPage.error)
        .where(
            CrawlPage.extraction_run_id == run_id,
            CrawlPage.state == CrawlPageState.FAILED,
        )
        .limit(50)
    )).all()
    failed_detail = [
        BlockedPageDetail(
            url=row.normalized_url or "",
            block_reason=row.block_reason or "FAILED",
            error=row.error,
        )
        for row in failed_rows
    ]

    return ExtractionProgress(
        crawl_pages_total=int(crawl_pages_total or 0),
        crawl_pages_pending=page_counts.get("PENDING", 0),
        crawl_pages_fetching=page_counts.get("FETCHING", 0),
        crawl_pages_extracted=page_counts.get("EXTRACTED", 0),
        crawl_pages_blocked=page_counts.get("BLOCKED", 0),
        crawl_pages_failed=page_counts.get("FAILED", 0),
        extracted_records_total=int(records or 0),
        exports_total=int(exports or 0),
        blocked_pages_detail=blocked_detail,
        failed_pages_detail=failed_detail,
    )


def _spec_response(spec: ExtractionSpec | None) -> ExtractionSpecResponse | None:
    if spec is None:
        return None
    return ExtractionSpecResponse(
        id=spec.id,
        project_id=spec.project_id,
        mode=spec.mode.value,
        fields=spec.fields or [],
        content_config=spec.content_config or {},
        url_patterns=spec.url_patterns or [],
        page_limit=spec.page_limit,
        export_format=spec.export_format,
        crawl_scope=spec.crawl_scope,
        interaction_profile=spec.interaction_profile or {},
        quality_summary=spec.quality_summary,
        created_at=spec.created_at,
        updated_at=spec.updated_at,
    )


def _frontier_preview_response(preview) -> FrontierPreviewResponse | None:
    if preview is None:
        return None
    return FrontierPreviewResponse(
        id=preview.id,
        project_id=preview.project_id,
        spec_id=preview.spec_id,
        scope_hash=preview.scope_hash,
        included_urls=[FrontierUrlDecision(**d) for d in (preview.included_urls or [])],
        excluded_urls=[FrontierUrlDecision(**d) for d in (preview.excluded_urls or [])],
        estimated_page_count=preview.estimated_page_count,
        warnings=preview.warnings or [],
        quality_summary=preview.quality_summary or {},
        created_at=preview.created_at,
    )


def _preview_response(preview) -> PreviewResponse | None:
    if preview is None:
        return None
    return PreviewResponse(
        id=preview.id,
        project_id=preview.project_id,
        spec_id=preview.spec_id,
        sample_records=preview.sample_records or [],
        warnings=preview.warnings or [],
        missing_fields=preview.missing_fields or [],
        quality_summary=preview.quality_summary or {},
        created_at=preview.created_at,
    )


def _extraction_quality(spec: ExtractionSpec | None) -> ExtractionQuality | None:
    if spec is None or not spec.quality_summary:
        return None
    qs = spec.quality_summary
    return ExtractionQuality(
        overall=qs.get("overall", "unknown"),
        field_success_rates=qs.get("field_success_rates", {}),
        missing_field_rates=qs.get("missing_field_rates", {}),
        warnings=qs.get("warnings", []),
    )


async def _project_response(db: AsyncSession, project: Project) -> ProjectResponse:
    spec = await ensure_default_spec(db, project)
    preview = await latest_preview(db, project.id)
    frontier_preview = await latest_frontier_preview(db, project.id)
    status_info = product_status_for(project)
    last_activity = project.updated_at or project.created_at
    # Compute preview_stale: spec was updated after the last preview ran.
    preview_stale = False
    if spec is not None and preview is not None:
        spec_updated = spec.updated_at or spec.created_at
        preview_created = preview.created_at
        if spec_updated is not None and preview_created is not None:
            preview_stale = spec_updated > preview_created
    return ProjectResponse(
        id=project.id,
        url=project.url,
        system_state=project.state.value,
        product_status=status_info.code,
        product_status_label=status_info.label,
        product_status_tone=status_info.tone,
        detected_type=detected_type(project),
        confidence=project.confidence,
        confidence_label=confidence_label(project.confidence, project.warnings or []),
        selected_field_count=selected_field_count(spec),
        extraction_mode=project.extraction_mode.value,
        workflow_mode=project.workflow_mode.value,
        render_mode=project.render_mode.value,
        provider_config_id=project.provider_config_id,
        browser_session_id=project.browser_session_id,
        warnings=project.warnings or [],
        analysis=project.analysis,
        fetch_metadata=project.fetch_metadata,
        spec=_spec_response(spec),
        preview=_preview_response(preview),
        frontier_preview=_frontier_preview_response(frontier_preview),
        extraction_quality=_extraction_quality(spec),
        preview_stale=preview_stale,
        current_extraction_run_id=project.current_extraction_run_id,
        progress=await _progress(db, project.id),
        last_activity=last_activity,
        error=project.error,
        error_code=project.error_code,
        created_at=project.created_at,
        updated_at=project.updated_at,
    )


async def _owned_project(db: AsyncSession, user: User, project_id: int) -> Project:
    project = await db.get(Project, project_id)
    if not project or project.user_id != user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    return project


@router.post(
    "/analyze",
    response_model=ProjectResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Analyze a URL and create a project",
)
@limiter.limit(SCRAPE_RATE_LIMIT)
async def analyze_project(
    request: Request,
    payload: ProjectAnalyzeRequest,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ProjectResponse:
    advanced = payload.advanced
    result = await admit_job(
        user=user,
        url=str(payload.url),
        extraction_mode=resolve_extraction_mode(
            str(payload.url),
            advanced.extraction_mode if advanced else None,
        ),
        workflow_mode=advanced.workflow_mode if advanced and advanced.workflow_mode else "GUIDED",
        render_mode=advanced.render_mode if advanced and advanced.render_mode else "AUTO",
        provider_config_id=advanced.provider_config_id if advanced else None,
        db=db,
    )

    if isinstance(result, JobAdmissionError):
        if result.error_type in {
            JobAdmissionErrorType.NO_PROVIDER_CONFIGURED,
            JobAdmissionErrorType.ACTIVE_JOB_LIMIT_REACHED,
        }:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={"message": result.message, "error_code": result.error_type.value},
            )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.message)

    background_tasks.add_task(
        execute_job_pipeline,
        job_id=result.job.id,
        provider_config_id=result.provider_config.id,
    )
    return await _project_response(db, result.job)


@router.get("", response_model=list[ProjectListItem], summary="List projects")
async def list_projects(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
) -> list[ProjectListItem]:
    result = await db.execute(
        select(Project)
        .where(Project.user_id == user.id)
        .order_by(Project.created_at.desc())
        .limit(limit)
        .offset(skip)
    )
    projects = result.scalars().all()

    # Batch-load the latest spec for each project in a single query.
    # The list view only needs selected_field_count, so we never auto-create
    # a spec here — that happens lazily when the user opens the workspace.
    specs_by_project: dict[int, ExtractionSpec | None] = {p.id: None for p in projects}
    if projects:
        spec_rows = await db.execute(
            select(ExtractionSpec)
            .where(ExtractionSpec.project_id.in_([p.id for p in projects]))
            .order_by(ExtractionSpec.created_at.desc(), ExtractionSpec.id.desc())
        )
        for spec in spec_rows.scalars().all():
            if specs_by_project.get(spec.project_id) is None:
                specs_by_project[spec.project_id] = spec

    items: list[ProjectListItem] = []
    for project in projects:
        spec = specs_by_project.get(project.id)
        status_info = product_status_for(project)
        items.append(
            ProjectListItem(
                id=project.id,
                url=project.url,
                system_state=project.state.value,
                product_status=status_info.code,
                product_status_label=status_info.label,
                product_status_tone=status_info.tone,
                detected_type=detected_type(project),
                confidence=project.confidence,
                confidence_label=confidence_label(project.confidence, project.warnings or []),
                selected_field_count=selected_field_count(spec),
                extraction_mode=project.extraction_mode.value,
                last_activity=project.updated_at or project.created_at,
                error=project.error,
                error_code=project.error_code,
            )
        )
    return items


@router.get("/{project_id}", response_model=ProjectResponse, summary="Get project detail")
async def get_project(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ProjectResponse:
    project = await _owned_project(db, user, project_id)
    response = await _project_response(db, project)
    await db.commit()
    return response


@router.patch("/{project_id}/spec", response_model=ExtractionSpecResponse, summary="Update extraction spec")
async def update_project_spec(
    project_id: int,
    payload: ExtractionSpecUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ExtractionSpecResponse:
    project = await _owned_project(db, user, project_id)
    spec = await ensure_default_spec(db, project)
    if spec is None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Project analysis is not ready")

    if payload.fields is not None:
        spec.fields = [field.model_dump() for field in payload.fields]
        invalid = flag_invalid_field_selectors(spec.fields)
        if invalid:
            logger.warning(
                "spec.invalid_selectors_on_save",
                extra={"project_id": project.id, "invalid_count": invalid},
            )
    if payload.content_config is not None:
        spec.content_config = payload.content_config
    if payload.url_patterns is not None:
        spec.url_patterns = payload.url_patterns
    if payload.page_limit is not None:
        spec.page_limit = payload.page_limit
    if payload.export_format is not None:
        spec.export_format = payload.export_format
    if payload.crawl_scope is not None:
        spec.crawl_scope = payload.crawl_scope.model_dump(mode="json")
    if payload.interaction_profile is not None:
        spec.interaction_profile = payload.interaction_profile.model_dump(mode="json")

    await db.commit()
    await db.refresh(spec)
    return _spec_response(spec)  # type: ignore[return-value]


@router.post(
    "/{project_id}/interactions/detect",
    response_model=ExtractionSpecResponse,
    summary="Detect page-variant controls",
)
async def detect_interactions(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ExtractionSpecResponse:
    """Scan the seed page for variant toggles (metric/imperial, per-100g/serving,
    …) and persist a disabled draft ``interaction_profile`` on the spec.

    Returns the updated spec (no raw page HTML). Detection runs against the
    static HTML, so it works whether or not a browser backend is installed; the
    proposed groups are interactive, so selecting one without a browser later
    fails with INTERACTION_BROWSER_REQUIRED (no silent downgrade)."""
    project = await _owned_project(db, user, project_id)
    spec = await ensure_default_spec(db, project)
    if spec is None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Project analysis is not ready",
        )

    try:
        validated_url = validate_url(project.normalized_url or project.url)
    except URLValidationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    effective_render_mode = project.render_mode.value
    if (
        effective_render_mode == "AUTO"
        and isinstance(project.fetch_metadata, dict)
        and project.fetch_metadata.get("render_mode_used") == "BROWSER"
    ):
        effective_render_mode = "BROWSER"

    session_cookies = None
    if project.browser_session_id is not None:
        session_cookies = await get_cookies_for_session(
            db, project.browser_session_id, owner_user_id=project.user_id
        )

    try:
        fetched = await fetch_url(
            validated_url, effective_render_mode, browser_session_cookies=session_cookies
        )
    except FetchError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    profile, new_fields = detect_interaction_profile(fetched.html, spec.fields or [])
    spec.interaction_profile = profile
    if new_fields is not None:
        # Numbered parallel columns (e.g. "Calories 1/2") were collapsed into
        # base fields the deterministic group overrides per variant.
        spec.fields = new_fields
    await db.commit()
    await db.refresh(spec)
    logger.info(
        "interaction.detected",
        extra={
            "project_id": project.id,
            "group_count": len(profile.get("groups") or []),
            "collapsed_fields": new_fields is not None,
        },
    )
    return _spec_response(spec)  # type: ignore[return-value]


@router.post("/{project_id}/preview", response_model=PreviewResponse, summary="Preview selected fields")
async def preview_project(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PreviewResponse:
    project = await _owned_project(db, user, project_id)
    spec = await latest_spec(db, project.id)
    if spec is None:
        spec = await ensure_default_spec(db, project)
    if spec is None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Select fields before previewing")
    if project.state not in {
        ProjectState.AWAITING_SETUP,
        ProjectState.ANALYSIS_READY,
        ProjectState.PREVIEW_READY,
    }:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Project is not ready for preview")

    try:
        preview = await create_preview(db, project, spec)
    except Exception as exc:
        await db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    await db.commit()
    await db.refresh(preview)
    return _preview_response(preview)  # type: ignore[return-value]


@router.post(
    "/{project_id}/frontier-preview",
    response_model=FrontierPreviewResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Generate crawl frontier preview",
)
async def create_frontier_preview_endpoint(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> FrontierPreviewResponse:
    project = await _owned_project(db, user, project_id)
    if project.state not in {
        ProjectState.AWAITING_SETUP,
        ProjectState.ANALYSIS_READY,
        ProjectState.PREVIEW_READY,
        ProjectState.COMPLETED,
    }:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Project is not ready for frontier preview")
    try:
        preview = await create_frontier_preview(db, project)
    except Exception as exc:
        await db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    if preview is None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Could not generate frontier preview: no spec or seed URL available",
        )
    await db.commit()
    await db.refresh(preview)
    return _frontier_preview_response(preview)  # type: ignore[return-value]


@router.get(
    "/{project_id}/frontier-preview",
    response_model=FrontierPreviewResponse,
    summary="Get latest crawl frontier preview",
)
async def get_frontier_preview(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> FrontierPreviewResponse:
    await _owned_project(db, user, project_id)
    preview = await latest_frontier_preview(db, project_id)
    if preview is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No frontier preview found")
    return _frontier_preview_response(preview)  # type: ignore[return-value]


@router.post("/{project_id}/extract", response_model=ProjectResponse, summary="Extract records")
async def extract_project(
    project_id: int,
    background_tasks: BackgroundTasks,
    payload: ExtractRequest | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ProjectResponse:
    project = await _owned_project(db, user, project_id)
    spec = await latest_spec(db, project.id)
    if spec is None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Select fields before extracting")
    preview = await latest_preview(db, project.id)
    extract_anyway = bool(payload.extract_anyway) if payload else False

    # Is the latest preview stale relative to the current spec?
    preview_is_stale = False
    if preview is not None:
        spec_updated = spec.updated_at or spec.created_at
        preview_created = preview.created_at
        preview_is_stale = bool(
            spec_updated is not None
            and preview_created is not None
            and spec_updated > preview_created
        )

    # Soft gates (bypassable with extract_anyway): no preview, or a stale one.
    if preview is None and not extract_anyway:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "message": "Preview before extracting, or choose extract anyway",
                "error_code": "NO_PREVIEW",
            },
        )
    if preview is not None and preview_is_stale and not extract_anyway:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "message": (
                    "Extraction spec was updated after the last preview. "
                    "Run preview again to verify selectors, or choose "
                    "extract anyway."
                ),
                "error_code": "STALE_PREVIEW",
            },
        )
    # Hard gate (NOT bypassable): a current preview that found zero structured
    # records means the selectors match nothing on the seed page, so extraction
    # would certainly produce zero rows. Forcing it wastes a crawl — require the
    # user to fix fields and re-preview. (A stale preview is handled above; once
    # bypassed we don't apply this, since the stale result may be outdated.)
    if (
        preview is not None
        and not preview_is_stale
        and spec.mode == ExtractionMode.STRUCTURED
        and len(preview.sample_records or []) == 0
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "message": (
                    "Preview found no records — the selectors do not match this "
                    "page. Adjust the fields and run Preview again before extracting."
                ),
                "error_code": "ZERO_PREVIEW_RECORDS",
            },
        )
    if project.state not in {
        ProjectState.AWAITING_SETUP,
        ProjectState.ANALYSIS_READY,
        ProjectState.PREVIEW_READY,
        ProjectState.COMPLETED,
    }:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Project is not ready for extraction")

    try:
        run = await start_project_extraction(db, project, spec)
    except ScopeConfirmationError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"message": str(exc), "error_code": exc.code, "scope": exc.scope},
        )
    except ExtractionAlreadyRunningError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"message": str(exc), "error_code": exc.code},
        )
    await db.commit()
    await db.refresh(project)
    background_tasks.add_task(execute_project_extraction, project.id, spec.id, run.id)
    return await _project_response(db, project)


@router.get(
    "/{project_id}/records",
    response_model=list[RecordResponse],
    summary="List extracted records (deprecated — use /records-page)",
    deprecated=True,
)
async def get_project_records(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=500),
) -> list[RecordResponse]:
    await _owned_project(db, user, project_id)
    records = await list_records(db, project_id, skip, limit)
    return [
        RecordResponse(
            id=record.id,
            source_url=record.source_url,
            raw_data=record.raw_data,
            normalized_data=record.normalized_data,
            warnings=record.warnings or [],
            created_at=record.created_at,
        )
        for record in records
    ]


@router.get(
    "/{project_id}/records-page",
    response_model=RecordPageResponse,
    summary="Paginated extracted records",
)
async def get_project_records_page(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
) -> RecordPageResponse:
    await _owned_project(db, user, project_id)
    total = await count_records(db, project_id)
    records = await list_records(db, project_id, skip, limit)
    items = [
        RecordResponse(
            id=record.id,
            source_url=record.source_url,
            raw_data=record.raw_data,
            normalized_data=record.normalized_data,
            warnings=record.warnings or [],
            created_at=record.created_at,
        )
        for record in records
    ]
    columns: list[str] = sorted({
        key
        for record in records
        for key in (record.normalized_data or record.raw_data or {}).keys()
    })
    has_more = (skip + limit) < total
    next_skip = skip + limit if has_more else None
    return RecordPageResponse(
        items=items,
        total=total,
        skip=skip,
        limit=limit,
        next_skip=next_skip,
        has_more=has_more,
        columns=columns,
    )


@router.get("/{project_id}/export", summary="Export generated results")
async def export_project(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    format: str = Query(default="csv", pattern="^(csv|json|xlsx)$"),
) -> Response:
    logger.info(
        "export.started",
        extra={"project_id": project_id, "user_id": user.id, "format": format},
    )
    start_time = time.monotonic()
    try:
        await _owned_project(db, user, project_id)
        total = await count_records(db, project_id)
        if total > 10_000:
            logger.warning(
                "export.large_export",
                extra={
                    "project_id": project_id,
                    "total_records": total,
                    "format": format,
                },
            )
        # Fetch all records in chunks to avoid silent truncation.
        # Previous implementation had a hard cap of 5000 that silently
        # dropped records beyond that limit.
        chunk_size = 1000
        data: list[dict[str, Any]] = []
        for skip in range(0, total, chunk_size):
            records = await list_records(db, project_id, skip, chunk_size)
            data.extend(record.normalized_data or record.raw_data for record in records)
        duration_ms = round((time.monotonic() - start_time) * 1000, 1)
        logger.info(
            "export.completed",
            extra={
                "project_id": project_id,
                "format": format,
                "record_count": len(data),
                "total_records": total,
                "duration_ms": duration_ms,
            },
        )
        # Resolve spec-defined field order: user_label priority matches the extractor's _field_key.
        spec = await latest_spec(db, project_id)
        field_order = _spec_field_order(spec)

        if format == "json":
            return Response(
                content=json.dumps(data),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="project-{project_id}.json"'},
            )
        if format == "xlsx":
            return Response(
                content=_xlsx_bytes(data, field_order=field_order),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="project-{project_id}.xlsx"'},
            )

        output = io.StringIO()
        fieldnames = _ordered_columns(data, field_order)
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="project-{project_id}.csv"'},
        )
    except Exception as exc:
        logger.error(
            "export.failed",
            extra={
                "project_id": project_id,
                "format": format,
                "error_type": type(exc).__name__,
            },
        )
        raise


def _spec_field_order(spec: Any | None) -> list[str]:
    """Return the user-visible field labels in spec order (user_label > label > name)."""
    if spec is None:
        return []
    out: list[str] = []
    for field in spec.fields or []:
        if not isinstance(field, dict) or not field.get("selected", True):
            continue
        key = field.get("user_label") or field.get("label") or field.get("name")
        if key:
            out.append(str(key))
    # Variant metadata columns sit right after the spec fields (in group order),
    # before any other extras and before source_url (added last by callers).
    # In merge mode the variants are encoded as per-variant column names instead,
    # so the fixed metadata columns are not emitted.
    profile = getattr(spec, "interaction_profile", None)
    if _interactions_enabled(profile) and not _interactions_merge_enabled(profile):
        for col in _interaction_metadata_columns(profile):
            if col not in out:
                out.append(col)
    return out


def _ordered_columns(rows: list[dict], field_order: list[str]) -> list[str]:
    """Merge spec field order with any extra keys found in the data.

    Priority: spec fields first (in spec order), then extras alphabetically,
    with ``source_url`` always last so the data columns are front-and-centre.
    """
    all_keys: set[str] = {key for row in rows for key in row.keys()}
    ordered = [k for k in field_order if k in all_keys]
    ordered_set = set(ordered)
    extras = sorted(k for k in all_keys if k not in ordered_set and k != "source_url")
    if "source_url" in all_keys:
        extras.append("source_url")
    return ordered + extras


_HEADER_FILL = PatternFill("solid", fgColor="1F618D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="top", wrap_text=False)
_ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
_THIN = Side(border_style="thin", color="D0D3D4")
_CELL_BORDER = Border(left=_THIN, right=_THIN, bottom=_THIN)


def _xlsx_bytes(rows: list[dict], *, field_order: list[str] | None = None) -> bytes:
    """Generate a styled XLSX workbook with openpyxl."""
    columns = _ordered_columns(rows, field_order or [])
    if not columns:
        columns = sorted({key for row in rows for key in row.keys()})

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _CELL_BORDER
    ws.row_dimensions[1].height = 28

    # Data rows with alternating row shading
    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            if value is None:
                value = ""
            elif not isinstance(value, (int, float, bool)):
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _DATA_ALIGN
            cell.border = _CELL_BORDER
            if fill:
                cell.fill = fill

    # Auto-fit column widths (capped to keep the sheet readable)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            val = row.get(col_name)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


@router.post("/{project_id}/cancel", response_model=ProjectResponse, summary="Cancel active project")
async def cancel_project(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ProjectResponse:
    project = await _owned_project(db, user, project_id)
    if project.state not in ACTIVE_PROJECT_STATES:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Project is not active")
    result = await transition_job_to_canceled(project_id, expected_states=ACTIVE_PROJECT_STATES)
    if not result.success or not result.job:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=result.error or "Cancellation failed")
    await record_project_event(
        project_id, user.id, "project.canceled", level="warning",
        message="Project canceled by user.",
    )
    return await _project_response(db, result.job)


@router.get(
    "/{project_id}/events",
    response_model=list[ProjectEventResponse],
    summary="Project activity log",
)
async def project_events(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    limit: int = Query(default=100, ge=1, le=500),
) -> list[ProjectEventResponse]:
    # Owner-checked; 404 on mismatch (do not reveal existence).
    await _owned_project(db, user, project_id)
    events = await list_project_events(db, project_id, user.id, limit=limit)
    return [ProjectEventResponse.from_event(event) for event in events]


@router.post("/{project_id}/retry", response_model=ProjectResponse, summary="Retry a failed project")
async def retry_project(
    project_id: int,
    background_tasks: BackgroundTasks,
    payload: RetryRequest | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ProjectResponse:
    project = await _owned_project(db, user, project_id)
    provider_override = payload.provider_config_id if payload else None
    try:
        project, provider = await retry_failed_project(
            db, project, user, provider_config_id=provider_override
        )
    except RetryError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"message": str(exc), "error_code": exc.error_code},
        ) from exc
    await db.commit()
    # Refresh so server-onupdate columns (e.g. updated_at) are not left expired,
    # which would trigger a sync lazy-load -> MissingGreenlet inside the response
    # builder. Build the response BEFORE queuing the task so a response failure
    # cannot silently drop the analysis task.
    await db.refresh(project)
    response = await _project_response(db, project)
    await record_project_event(
        project.id, user.id, "project.retried",
        message="Project retry requested by user.",
    )
    if provider is not None:
        background_tasks.add_task(
            execute_job_pipeline,
            job_id=project.id,
            provider_config_id=provider.id,
        )
    return response


@router.patch(
    "/{project_id}/session",
    response_model=ProjectResponse,
    summary="Assign or clear a browser session for a project",
)
async def set_project_session(
    project_id: int,
    browser_session_id: int | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ProjectResponse:
    from app.services.session_service import get_session  # noqa: PLC0415
    project = await _owned_project(db, user, project_id)
    if browser_session_id is not None:
        session = await get_session(db, user, browser_session_id)
        if session is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Session not found",
            )
    project.browser_session_id = browser_session_id
    await db.commit()
    # Refresh to reload server-onupdate columns before building the response
    # (same MissingGreenlet hazard as retry).
    await db.refresh(project)
    return await _project_response(db, project)


@router.delete(
    "/{project_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_model=None,
    response_class=Response,
    summary="Delete project",
)
async def delete_project(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Response:
    project = await _owned_project(db, user, project_id)
    if project.state not in DELETABLE_PROJECT_STATES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete an active project",
        )
    try:
        await delete_project_tree(db, project)
    except Exception:
        logger.exception(
            "project.delete_failed",
            extra={"project_id": project_id, "user_id": user.id},
        )
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                "Failed to delete project. A background task may hold a row "
                "lock; wait 30 seconds and try again."
            ),
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)
