"""Layer A - live, service-level E2E for the whole scraping pipeline.

Drives the REAL services end to end against REAL websites, across every crawl
scope mode and both extraction modes, plus the variant/interaction path and the
export column logic:

    fetch_url -> normalize_crawl_scope / recommend_scope
              -> derive_include_patterns_from_links
              -> classify_links_for_scope / discover_links_for_scope  (BFS crawl)
              -> extract_records_from_html / extract_records_with_variants
              -> _spec_field_order / _ordered_columns -> CSV/JSON/XLSX bytes

No DB and no LLM are involved: specs are constructed deterministically so every
assertion is about the PIPELINE's own correctness, not model quality. This is
the layer that caught the per-serving bug.

Each scenario is classified honestly:
    PASS   - all assertions held
    FAIL   - a real, generic pipeline bug (must be fixed)
    LIMIT  - a real but acceptable limitation (reported, not "fixed" with a hack)
    ENV    - the real site/browser was unreachable/flaky (not a pipeline bug)

Run:
    venv\\Scripts\\python.exe -m tests.e2e.run_live_pipeline            # all sites
    venv\\Scripts\\python.exe -m tests.e2e.run_live_pipeline books      # one site
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import sys
import time
from dataclasses import dataclass, field as dc_field
from types import SimpleNamespace
from typing import Any, Callable

from openpyxl import load_workbook

from app.api.v1.endpoints.projects import (
    _ordered_columns,
    _spec_field_order,
    _xlsx_bytes,
)
from app.core.logging_config import configure_logging
from app.models.job import ExtractionMode
from app.services.crawl_scope import (
    derive_include_patterns_from_links,
    discover_links_for_scope,
    normalize_crawl_scope,
)
from app.services.extractor import extract_records_from_html
from app.services.fetcher import FetchError, apply_interactions_and_capture, fetch_url
from app.services.interaction_detect import detect_interaction_profile
from app.services.interaction_extraction import extract_records_with_variants

configure_logging()
logger = logging.getLogger("e2e.live")

# Politeness + resilience knobs (real sites).
FETCH_TRIES = 3
FETCH_RETRY_DELAY = 2.0
CRAWL_DELAY = 0.8


# --------------------------------------------------------------------------- #
# small builders
# --------------------------------------------------------------------------- #
def f(name: str, selector: str, type_: str = "string") -> dict[str, Any]:
    return {
        "name": name,
        "label": name,
        "user_label": name,
        "selector": selector,
        "type": type_,
        "selected": True,
    }


def make_spec(
    *,
    mode: ExtractionMode,
    fields: list[dict[str, Any]] | None = None,
    content_config: dict[str, Any] | None = None,
    interaction_profile: dict[str, Any] | None = None,
    crawl_scope: dict[str, Any] | None = None,
    page_limit: int = 6,
    export_format: str = "csv",
) -> SimpleNamespace:
    return SimpleNamespace(
        mode=mode,
        fields=fields or [],
        content_config=content_config or {},
        interaction_profile=interaction_profile or {},
        crawl_scope=crawl_scope,
        page_limit=page_limit,
        export_format=export_format,
    )


async def fetch_with_retry(
    url: str, render: str = "AUTO", cookies: list[dict] | None = None
) -> Any:
    """Fetch with bounded retries so transient env flakiness != pipeline bug."""
    last: Exception | None = None
    for attempt in range(1, FETCH_TRIES + 1):
        try:
            return await fetch_url(url, render, browser_session_cookies=cookies)
        except FetchError as exc:
            last = exc
            logger.warning(
                "fetch attempt %s/%s failed url=%s code=%s",
                attempt, FETCH_TRIES, url, exc.error_code,
            )
            if attempt < FETCH_TRIES:
                await asyncio.sleep(FETCH_RETRY_DELAY)
    assert last is not None
    raise last


# --------------------------------------------------------------------------- #
# generic BFS crawl that mirrors project_extraction's frontier semantics
# --------------------------------------------------------------------------- #
@dataclass
class CrawlResult:
    records: list[Any]
    pages: list[dict[str, Any]]      # {url, ok, records|error}
    scope: dict[str, Any]
    derived_patterns: list[str] | None
    env_error: str | None = None

    @property
    def ok_pages(self) -> list[dict[str, Any]]:
        return [p for p in self.pages if p.get("ok")]

    @property
    def source_urls(self) -> list[str]:
        return sorted({str(r.normalized_data.get("source_url")) for r in self.records})


async def crawl_and_extract(
    *,
    seed_url: str,
    scope_mode: str,
    spec: SimpleNamespace,
    analysis: dict[str, Any],
    render: str = "AUTO",
) -> CrawlResult:
    """Bounded BFS that exercises the real scope + extraction services."""
    page_limit = int(spec.page_limit)
    project = SimpleNamespace(url=seed_url, normalized_url=seed_url, analysis=analysis)

    scope = normalize_crawl_scope(
        {"mode": scope_mode, **(spec.crawl_scope or {})},
        seed_url=seed_url,
        page_limit=page_limit,
    )

    try:
        seed_fetch = await fetch_with_retry(seed_url, render)
    except FetchError as exc:
        return CrawlResult([], [{"url": seed_url, "ok": False, "error": exc.error_code}],
                           scope, None, env_error=exc.error_code)

    root = seed_fetch.final_url
    # Self-configure include patterns from the seed's real links (COLLECTION/DATASET).
    derived = derive_include_patterns_from_links(
        scope, html=seed_fetch.html, seed_url=root, analysis=analysis
    )
    if derived:
        scope = {**scope, "include_patterns": derived}

    visited: set[str] = {root}
    queue: list[tuple[str, int, Any]] = [(root, 0, seed_fetch)]
    records: list[Any] = []
    pages: list[dict[str, Any]] = []

    while queue and len(pages) < page_limit:
        url, depth, prefetched = queue.pop(0)
        fetched = prefetched
        if fetched is None:
            await asyncio.sleep(CRAWL_DELAY)
            try:
                fetched = await fetch_with_retry(url, render)
            except FetchError as exc:
                pages.append({"url": url, "ok": False, "error": exc.error_code})
                continue

        recs = extract_records_from_html(
            fetched.html, source_url=fetched.final_url, project=project,
            spec=spec, max_records=1000,
        )
        records.extend(recs)
        pages.append({"url": fetched.final_url, "ok": True, "records": len(recs),
                      "depth": depth})

        remaining = page_limit - len(pages)
        if remaining > 0:
            children = discover_links_for_scope(
                fetched.html, page_url=fetched.final_url, root_url=root,
                scope=scope, analysis=analysis, limit=max(remaining * 4, 8),
                source_depth=depth,
            )
            for child in children:
                if child not in visited and len(visited) < page_limit * 6:
                    visited.add(child)
                    queue.append((child, depth + 1, None))

    return CrawlResult(records, pages, scope, derived)


# --------------------------------------------------------------------------- #
# export validation (mode-agnostic, pure python)
# --------------------------------------------------------------------------- #
def validate_export(spec: SimpleNamespace, records: list[Any]) -> list[str]:
    """Round-trip CSV/JSON/XLSX through the REAL export column logic."""
    fails: list[str] = []
    rows = [dict(r.normalized_data) for r in records]
    if not rows:
        return ["export: no rows to export"]
    field_order = _spec_field_order(spec)
    cols = _ordered_columns(rows, field_order)
    if not cols:
        fails.append("export: _ordered_columns returned no columns")
        return fails
    if "source_url" in {k for row in rows for k in row} and cols[-1] != "source_url":
        fails.append(f"export: source_url not last column (got {cols[-1]!r})")
    # spec fields must lead the column order
    for i, label in enumerate([c for c in field_order if c in cols]):
        if cols[i] != label:
            fails.append(f"export: spec field order broken at {i}: {cols[i]!r} != {label!r}")
            break

    # CSV bytes
    try:
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in cols})
        parsed = list(csv.DictReader(io.StringIO(buf.getvalue())))
        if len(parsed) != len(rows):
            fails.append(f"export CSV: round-trip row count {len(parsed)} != {len(rows)}")
    except Exception as exc:  # noqa: BLE001
        fails.append(f"export CSV raised: {exc}")

    # JSON bytes
    try:
        blob = json.dumps(rows, ensure_ascii=False, default=str)
        if len(json.loads(blob)) != len(rows):
            fails.append("export JSON: round-trip row count mismatch")
    except Exception as exc:  # noqa: BLE001
        fails.append(f"export JSON raised: {exc}")

    # XLSX bytes
    try:
        data = _xlsx_bytes(rows, field_order=field_order)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        if header != cols:
            fails.append(f"export XLSX: header {header} != columns {cols}")
        if ws.max_row != len(rows) + 1:
            fails.append(f"export XLSX: {ws.max_row - 1} data rows != {len(rows)}")
    except Exception as exc:  # noqa: BLE001
        fails.append(f"export XLSX raised: {exc}")

    return fails


# --------------------------------------------------------------------------- #
# scenario model
# --------------------------------------------------------------------------- #
@dataclass
class Scenario:
    site: str
    name: str
    seed: str
    ext_mode: ExtractionMode
    scope_mode: str
    analysis: dict[str, Any]
    check: Callable[[CrawlResult], list[str]]
    fields: list[dict[str, Any]] = dc_field(default_factory=list)
    content_config: dict[str, Any] | None = None
    crawl_scope: dict[str, Any] | None = None
    page_limit: int = 6
    render: str = "AUTO"
    do_export: bool = True
    # Some honest, expected limitations: a non-empty failure list is downgraded
    # from FAIL to LIMIT when classify(failures) says so.
    limit_when: Callable[[list[str], CrawlResult], bool] | None = None


@dataclass
class Outcome:
    scenario: Scenario
    status: str          # PASS | FAIL | LIMIT | ENV
    failures: list[str]
    evidence: str
    seconds: float


def nonempty(records: list[Any], key: str) -> int:
    return sum(1 for r in records if str(r.normalized_data.get(key) or "").strip())


# --------------------------------------------------------------------------- #
# runner
# --------------------------------------------------------------------------- #
async def run_scenario(sc: Scenario) -> Outcome:
    start = time.monotonic()
    spec = make_spec(
        mode=sc.ext_mode,
        fields=sc.fields,
        content_config=sc.content_config,
        crawl_scope=sc.crawl_scope,
        page_limit=sc.page_limit,
    )
    res = await crawl_and_extract(
        seed_url=sc.seed, scope_mode=sc.scope_mode, spec=spec,
        analysis=sc.analysis, render=sc.render,
    )
    secs = time.monotonic() - start

    if res.env_error and not res.records:
        return Outcome(sc, "ENV", [f"seed unreachable: {res.env_error}"],
                       f"pages={res.pages}", secs)

    failures = sc.check(res)
    if sc.do_export and res.records and not any("export" in x for x in failures):
        failures += validate_export(spec, res.records)

    status = "PASS"
    if failures:
        status = "LIMIT" if (sc.limit_when and sc.limit_when(failures, res)) else "FAIL"

    evidence = (
        f"pages_ok={len(res.ok_pages)}/{len(res.pages)} "
        f"records={len(res.records)} src_urls={len(res.source_urls)} "
        f"derived={res.derived_patterns} mode={res.scope.get('mode')}"
    )
    return Outcome(sc, status, failures, evidence, secs)


async def run_variant_scenario() -> Outcome:
    """The calories.info serving_basis path - the headline correctness case.

    Mirrors tests/manual/verify_interaction_variants but asserts via
    (Food, serving_basis) and the collapsed field labels, using the real browser
    capture (camoufox -> chromium cascade)."""
    start = time.monotonic()
    url = "https://www.calories.info/food/beef-veal"
    row = "table.MuiTable-root tr"
    fails: list[str] = []
    try:
        fetched = await fetch_with_retry(url, "AUTO")
    except FetchError as exc:
        return Outcome(
            Scenario("calories", "variants serving_basis (CURRENT_PAGE)", url,
                     ExtractionMode.STRUCTURED, "CURRENT_PAGE", {}, lambda r: []),
            "ENV", [f"seed unreachable: {exc.error_code}"], "", time.monotonic() - start)

    ord_fields = [
        f("Food", "td:nth-child(1) p"),
        f("Serving Size (per 100 g)", "td:nth-child(2) p"),
        f("Calories (per 100 g)", "td:nth-child(3)", "number"),
        f("Serving Size (alternate column)", "td:nth-child(4) p"),
        f("Calories (alternate column)", "td:nth-child(5)", "number"),
    ]
    mprof, mfields = detect_interaction_profile(
        fetched.html, ord_fields, repeated_item_selector=row
    )
    mfields = mfields or ord_fields
    labels = [str(x.get("label") or x.get("name")) for x in mfields]
    if any("alt" in lb.lower() for lb in labels):
        fails.append(f"collapsed fields still expose an alt column: {labels}")
    execs = {g["metadata_key"]: g["execution"] for g in mprof["groups"]}
    if execs.get("serving_basis") != "mixed":
        fails.append(f"serving_basis not mixed: {execs}")

    async def cb(recipes: dict[str, list[dict]]) -> dict[str, str]:
        return await apply_interactions_and_capture(fetched.final_url, recipes)

    spec = make_spec(
        mode=ExtractionMode.STRUCTURED, fields=mfields,
        interaction_profile={**mprof, "enabled": True},
    )
    recs, _w = await extract_records_with_variants(
        base_html=fetched.html, source_url=fetched.final_url,
        project=SimpleNamespace(analysis={"repeated_item_selector": row}),
        spec=spec, max_records=1000, fetch_variant_htmls=cb,
    )
    skey = next((x.get("user_label") or x.get("label") or x.get("name")
                 for x in mfields if "serving" in str(x.get("label", "")).lower()),
                "Serving Size")
    bases = {str(r.normalized_data.get("serving_basis")) for r in recs}
    if len(bases) < 2:
        fails.append(f"expected >=2 serving_basis values, got {bases}")
    got = {(d.get("Food"), str(d.get("serving_basis"))): (d.get(skey), d.get("Calories"))
           for d in (r.normalized_data for r in recs)}
    serv, cal = got.get(("Beef", "Show per serving"), (None, None))
    if not (serv and "portion" in str(serv).lower()):
        fails.append(f"Beef per-serving serving size wrong: {serv!r} (want '...portion...')")
    elif str(cal) != "265":
        fails.append(f"Beef per-serving calories wrong: {cal!r} (want 265)")

    secs = time.monotonic() - start
    sc = Scenario("calories", "variants serving_basis (CURRENT_PAGE)", url,
                  ExtractionMode.STRUCTURED, "CURRENT_PAGE", {}, lambda r: [])
    status = "PASS" if not fails else "FAIL"
    evidence = (f"records={len(recs)} serving_basis={sorted(bases)} "
                f"Beef/per-serving={got.get(('Beef','Show per serving'))}")
    return Outcome(sc, status, fails, evidence, secs)


async def run_collection_variant_scenario() -> Outcome:
    """COLLECTION + browser variants on a NON-seed child page (/food/meat).

    Regression for project-189: the COLLECTION crawl applies the seed's variant
    spec to every sibling /food/* page. Under concurrent crawling the browser
    toggle silently degraded to the page's STATIC values on child pages, so the
    per-serving SERVING SIZE stayed '100 g' (calories were correct). Build the
    spec from the seed, confirm /food/meat is reachable via the sibling glob, then
    extract /food/meat with the real browser capture and assert the
    browser-rendered per-serving sizes Codex flagged."""
    start = time.monotonic()
    seed_url = "https://www.calories.info/food/beef-veal"
    meat_url = "https://www.calories.info/food/meat"
    row = "table.MuiTable-root tr"
    sc = Scenario("calories", "COLLECTION+variants /food/meat (child page)", meat_url,
                  ExtractionMode.STRUCTURED, "COLLECTION", {}, lambda r: [])
    fails: list[str] = []
    try:
        seed = await fetch_with_retry(seed_url, "AUTO")
    except FetchError as exc:
        return Outcome(sc, "ENV", [f"seed unreachable: {exc.error_code}"], "",
                       time.monotonic() - start)

    ord_fields = [
        f("Food", "td:nth-child(1) p"),
        f("Serving Size (per 100 g)", "td:nth-child(2) p"),
        f("Calories (per 100 g)", "td:nth-child(3)", "number"),
        f("Serving Size (alternate column)", "td:nth-child(4) p"),
        f("Calories (alternate column)", "td:nth-child(5)", "number"),
    ]
    mprof, mfields = detect_interaction_profile(seed.html, ord_fields, repeated_item_selector=row)
    mfields = mfields or ord_fields
    skey = next((x.get("user_label") or x.get("label") or x.get("name")
                 for x in mfields if "serving" in str(x.get("label", "")).lower()),
                "Serving Size")

    # /food/meat must be reachable via the COLLECTION sibling glob.
    scope = normalize_crawl_scope({"mode": "COLLECTION"}, seed_url=seed.final_url, page_limit=10)
    derived = derive_include_patterns_from_links(
        scope, html=seed.html, seed_url=seed.final_url,
        analysis={"detail_link_selector": "a[href*='/food/']"})
    if derived:
        scope = {**scope, "include_patterns": derived}
    siblings = discover_links_for_scope(
        seed.html, page_url=seed.final_url, root_url=seed.final_url,
        scope=scope, analysis=None, limit=200, source_depth=0)
    if not any("/food/meat" in u for u in siblings):
        fails.append(f"/food/meat not reachable via COLLECTION glob {derived}")

    meat = await fetch_with_retry(meat_url, "AUTO")

    async def cb(recipes: dict[str, list[dict]]) -> dict[str, str]:
        return await apply_interactions_and_capture(meat.final_url, recipes)

    spec = make_spec(mode=ExtractionMode.STRUCTURED, fields=mfields,
                     interaction_profile={**mprof, "enabled": True})
    recs, _w = await extract_records_with_variants(
        base_html=meat.html, source_url=meat.final_url,
        project=SimpleNamespace(analysis={"repeated_item_selector": row}),
        spec=spec, max_records=2000, fetch_variant_htmls=cb)
    got = {(d.get("Food"), str(d.get("serving_basis"))): (d.get(skey), d.get("Calories"))
           for d in (r.normalized_data for r in recs)}
    bserv, bcal = got.get(("Beef", "Show per serving"), (None, None))
    cserv, ccal = got.get(("Chicken", "Show per serving"), (None, None))
    if not (bserv and "portion" in str(bserv).lower() and str(bcal) == "265"):
        fails.append(f"/food/meat Beef per-serving wrong: serving={bserv!r} cal={bcal!r} "
                     f"(want '1 portion (170 g)'/265)")
    if not (cserv and "piece" in str(cserv).lower() and str(ccal) == "764"):
        fails.append(f"/food/meat Chicken per-serving wrong: serving={cserv!r} cal={ccal!r} "
                     f"(want '1/2 piece (460 g)'/764)")

    evidence = (f"derived={derived} Beef/per-serving=({bserv!r},{bcal!r}) "
                f"Chicken/per-serving=({cserv!r},{ccal!r})")
    return Outcome(sc, "PASS" if not fails else "FAIL", fails, evidence,
                   time.monotonic() - start)


# --------------------------------------------------------------------------- #
# scenarios
# --------------------------------------------------------------------------- #
BOOKS = "https://books.toscrape.com/"
BOOKS_ANALYSIS = {"repeated_item_selector": "article.product_pod",
                  "pagination_selector": "li.next a",
                  "detail_link_selector": "article.product_pod h3 a"}
BOOKS_DETAIL_ANALYSIS = {"repeated_item_selector": "div.product_main",
                         "detail_link_selector": "article.product_pod h3 a"}
BOOKS_LIST_FIELDS = [f("Title", "h3 a"), f("Price", "p.price_color"),
                     f("Availability", "p.instock.availability")]
BOOKS_DETAIL_FIELDS = [f("Title", "div.product_main h1"),
                       f("Price", "div.product_main p.price_color"),
                       f("Stock", "div.product_main p.availability")]

HN = "https://news.ycombinator.com/"
HN_ANALYSIS = {"repeated_item_selector": "tr.athing",
               "pagination_selector": "a.morelink"}
HN_FIELDS = [f("Title", "span.titleline a"), f("Rank", "span.rank")]

CAL = "https://www.calories.info/food/beef-veal"
CAL_ANALYSIS = {"repeated_item_selector": "table.MuiTable-root tr",
                "detail_link_selector": "a[href*='/food/']"}
CAL_FIELDS = [f("Food", "td:nth-child(1) p"),
              f("Serving Size", "td:nth-child(2) p"),
              f("Calories", "td:nth-child(3)", "number")]

# Documentation page (MkDocs Material): headline CONTENT-as-Markdown case.
FASTAPI = "https://fastapi.tiangolo.com/"
FASTAPI_ANALYSIS: dict[str, Any] = {}


def _is_category(u: str) -> bool:
    return "/category/" in u


def chk_books_current(r: CrawlResult) -> list[str]:
    fails = []
    if len(r.ok_pages) != 1:
        fails.append(f"CURRENT_PAGE crawled {len(r.ok_pages)} pages (want 1)")
    if not (15 <= len(r.records) <= 25):
        fails.append(f"expected ~20 books, got {len(r.records)}")
    if nonempty(r.records, "Title") != len(r.records):
        fails.append("some books have empty Title")
    if nonempty(r.records, "Price") != len(r.records):
        fails.append("some books have empty Price")
    return fails


def chk_books_pagination(r: CrawlResult) -> list[str]:
    fails = []
    if len(r.ok_pages) < 2:
        fails.append(f"PAGINATION followed only {len(r.ok_pages)} page(s)")
    if len(r.records) <= 20:
        fails.append(f"PAGINATION yielded {len(r.records)} records (want >20)")
    paged = [p for p in r.ok_pages if "page-" in p["url"]]
    if not paged:
        fails.append("no /catalogue/page-N.html page was followed")
    return fails


def chk_books_collection(r: CrawlResult) -> list[str]:
    # books home has no parent path, so sibling-glob derivation can't apply here;
    # this scenario probes that honestly (see limit_when below).
    fails = []
    if not r.derived_patterns:
        fails.append("COLLECTION derived no include_patterns from the seed")
    if len(r.source_urls) < 2:
        fails.append(f"COLLECTION crawled {len(r.source_urls)} distinct page(s)")
    return fails


def chk_books_dataset(r: CrawlResult) -> list[str]:
    fails = []
    detail_pages = [p for p in r.ok_pages
                    if not _is_category(p["url"]) and "/catalogue/" in p["url"]
                    and "page-" not in p["url"] and p["url"].rstrip("/") != BOOKS.rstrip("/")]
    if len(detail_pages) < 2:
        fails.append(f"DATASET crawled {len(detail_pages)} detail page(s) (want >=2)")
    if nonempty(r.records, "Title") < 2:
        fails.append("DATASET detail records missing Title")
    return fails


def chk_books_content(r: CrawlResult) -> list[str]:
    fails = []
    if len(r.records) != 1:
        fails.append(f"CONTENT produced {len(r.records)} records (want 1)")
        return fails
    text = str(r.records[0].normalized_data.get("content") or "")
    if len(text) < 200:
        fails.append(f"CONTENT too short ({len(text)} chars)")
    low = text.lower()
    if "product description" not in low and "upc" not in low:
        fails.append("CONTENT missing the book's description/details text")
    # chrome should be stripped
    if "<script" in text or "function(" in text:
        fails.append("CONTENT contains script chrome")
    return fails


def chk_books_full_site(r: CrawlResult) -> list[str]:
    fails = []
    if len(r.ok_pages) < 2:
        fails.append(f"FULL_SITE (bounded) followed only {len(r.ok_pages)} page(s)")
    return fails


def chk_cal_collection(r: CrawlResult) -> list[str]:
    fails = []
    if not r.derived_patterns or not any("/food/" in p for p in r.derived_patterns):
        fails.append(f"COLLECTION derived no /food/* pattern (got {r.derived_patterns})")
    if len(r.source_urls) < 2:
        fails.append(f"COLLECTION crawled {len(r.source_urls)} distinct /food page(s)")
    if not all("/food/" in u for u in r.source_urls):
        fails.append(f"COLLECTION strayed off /food/: {r.source_urls}")
    return fails


def chk_cal_content(r: CrawlResult) -> list[str]:
    fails = []
    if len(r.records) != 1:
        fails.append(f"CONTENT produced {len(r.records)} records (want 1)")
        return fails
    text = str(r.records[0].normalized_data.get("content") or "")
    if len(text) < 200:
        fails.append(f"CONTENT too short ({len(text)} chars)")
    return fails


def chk_hn_current(r: CrawlResult) -> list[str]:
    fails = []
    if not (20 <= len(r.records) <= 40):
        fails.append(f"HN front page: expected ~30 items, got {len(r.records)}")
    if nonempty(r.records, "Title") < 20:
        fails.append("HN items missing Title")
    return fails


def chk_hn_pagination(r: CrawlResult) -> list[str]:
    fails = []
    if len(r.ok_pages) < 2:
        fails.append(f"HN PAGINATION followed only {len(r.ok_pages)} page(s)")
    if not any("p=" in p["url"] or "p2" in p["url"] for p in r.ok_pages[1:]):
        fails.append("HN PAGINATION did not follow ?p=N")
    return fails


def chk_hn_content(r: CrawlResult) -> list[str]:
    fails = []
    text = str(r.records[0].normalized_data.get("content") or "") if r.records else ""
    if len(text) < 200:
        fails.append(f"HN CONTENT too short ({len(text)} chars)")
    return fails


def chk_fastapi_content(r: CrawlResult) -> list[str]:
    """The headline CONTENT-as-Markdown case on a real docs page: code blocks must
    survive as fenced blocks (NOT shredded one token per line), heading-anchor
    pilcrows must be gone, and at least one real code line must be intact."""
    fails = []
    if len(r.records) != 1:
        fails.append(f"CONTENT produced {len(r.records)} records (want 1)")
        return fails
    text = str(r.records[0].normalized_data.get("content") or "")
    if len(text) < 500:
        fails.append(f"CONTENT too short ({len(text)} chars)")
    if "```" not in text:
        fails.append("CONTENT has no fenced code block (Markdown not emitted)")
    if "¶" in text:
        fails.append("CONTENT still contains ¶ heading-anchor pilcrows")
    if "from fastapi import FastAPI" not in text:
        fails.append("CONTENT code block is shredded (token-per-line) or missing")
    if "#" not in text:
        fails.append("CONTENT has no Markdown headings")
    return fails


SCENARIOS: list[Scenario] = [
    # books.toscrape.com - the clean sandbox: exercises every mode + content
    Scenario("books", "CURRENT_PAGE structured", BOOKS, ExtractionMode.STRUCTURED,
             "CURRENT_PAGE", BOOKS_ANALYSIS, chk_books_current, BOOKS_LIST_FIELDS,
             page_limit=1),
    Scenario("books", "PAGINATION structured", BOOKS, ExtractionMode.STRUCTURED,
             "PAGINATION", BOOKS_ANALYSIS, chk_books_pagination, BOOKS_LIST_FIELDS,
             page_limit=4),
    Scenario("books", "COLLECTION structured (home, no parent)", BOOKS,
             ExtractionMode.STRUCTURED, "COLLECTION", BOOKS_ANALYSIS,
             chk_books_collection, BOOKS_LIST_FIELDS, page_limit=5,
             # home has no parent path -> sibling glob can't be derived; that is a
             # documented design property, not a bug.
             limit_when=lambda fa, r: not r.derived_patterns),
    Scenario("books", "DATASET listing+detail", BOOKS, ExtractionMode.STRUCTURED,
             "DATASET", BOOKS_DETAIL_ANALYSIS, chk_books_dataset, BOOKS_DETAIL_FIELDS,
             page_limit=6),
    Scenario("books", "CONTENT detail page",
             "https://books.toscrape.com/catalogue/a-light-in-the-attic_1000/index.html",
             ExtractionMode.CONTENT, "CURRENT_PAGE", BOOKS_ANALYSIS, chk_books_content,
             content_config={}, page_limit=1, do_export=False),
    Scenario("books", "FULL_SITE bounded", BOOKS, ExtractionMode.STRUCTURED,
             "FULL_SITE", BOOKS_ANALYSIS, chk_books_full_site, BOOKS_LIST_FIELDS,
             page_limit=5),

    # calories.info - COLLECTION sibling-glob + content (variants run separately)
    Scenario("calories", "COLLECTION /food/* siblings", CAL, ExtractionMode.STRUCTURED,
             "COLLECTION", CAL_ANALYSIS, chk_cal_collection, CAL_FIELDS, page_limit=4),
    Scenario("calories", "CONTENT current page", CAL, ExtractionMode.CONTENT,
             "CURRENT_PAGE", CAL_ANALYSIS, chk_cal_content, content_config={},
             page_limit=1, do_export=False),

    # news.ycombinator.com - real-world robustness
    Scenario("hn", "CURRENT_PAGE structured", HN, ExtractionMode.STRUCTURED,
             "CURRENT_PAGE", HN_ANALYSIS, chk_hn_current, HN_FIELDS, page_limit=1),
    Scenario("hn", "PAGINATION structured", HN, ExtractionMode.STRUCTURED,
             "PAGINATION", HN_ANALYSIS, chk_hn_pagination, HN_FIELDS, page_limit=3),
    Scenario("hn", "CONTENT current page", HN, ExtractionMode.CONTENT,
             "CURRENT_PAGE", HN_ANALYSIS, chk_hn_content, content_config={},
             page_limit=1, do_export=False),

    # fastapi.tiangolo.com - CONTENT-as-Markdown on a real documentation page
    Scenario("fastapi", "CONTENT markdown (docs page)", FASTAPI,
             ExtractionMode.CONTENT, "CURRENT_PAGE", FASTAPI_ANALYSIS,
             chk_fastapi_content, content_config={}, page_limit=1,
             do_export=False),
]


async def main() -> int:
    only = {a.lower() for a in sys.argv[1:]}
    scenarios = [s for s in SCENARIOS if not only or s.site in only]
    run_variants = (not only) or ("calories" in only) or ("variants" in only)

    outcomes: list[Outcome] = []
    for sc in scenarios:
        logger.info("=== %s :: %s ===", sc.site, sc.name)
        try:
            outcomes.append(await run_scenario(sc))
        except Exception as exc:  # noqa: BLE001
            logger.exception("scenario crashed")
            outcomes.append(Outcome(sc, "FAIL", [f"harness crash: {exc!r}"], "", 0.0))

    if run_variants:
        logger.info("=== calories :: variants serving_basis (browser) ===")
        try:
            outcomes.append(await run_variant_scenario())
        except Exception as exc:  # noqa: BLE001
            logger.exception("variant scenario crashed")
        logger.info("=== calories :: COLLECTION+variants /food/meat (child page) ===")
        try:
            outcomes.append(await run_collection_variant_scenario())
        except Exception as exc:  # noqa: BLE001
            logger.exception("collection variant scenario crashed")

    # report
    print("\n" + "=" * 78)
    print("LAYER A - LIVE PIPELINE E2E RESULTS")
    print("=" * 78)
    width = max(len(f"{o.scenario.site}/{o.scenario.name}") for o in outcomes)
    counts = {"PASS": 0, "FAIL": 0, "LIMIT": 0, "ENV": 0}
    for o in outcomes:
        counts[o.status] += 1
        tag = f"{o.scenario.site}/{o.scenario.name}".ljust(width)
        print(f"[{o.status:5}] {tag}  {o.seconds:5.1f}s  {o.evidence}")
        for fa in o.failures:
            print(f"          - {fa}")
    print("-" * 78)
    print(f"PASS={counts['PASS']} FAIL={counts['FAIL']} "
          f"LIMIT={counts['LIMIT']} ENV={counts['ENV']}")
    print("=" * 78)
    return counts["FAIL"]


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
