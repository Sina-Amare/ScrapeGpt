"""Layer B - HTTP API E2E for the full project workflow (real FastAPI stack).

Drives the REAL backend over httpx against a local fixture site, exercising the
endpoints + lifecycle that the seeded Phase-2.5 validation
(tests/validation/run_validation.py) does NOT actually run live:

  * live analyze (create project from a URL via the user's free provider)
  * preview (sample selectors)
  * frontier preview (POST/GET)
  * the LIVE extract -> COMPLETED path through the worker, then records-page
    + export CSV/JSON/XLSX on the records the worker actually produced
  * every extract gate code: NO_PREVIEW, STALE_PREVIEW, ZERO_PREVIEW_RECORDS,
    SCOPE_NOT_CONFIRMED
  * lifecycle: retry / cancel / delete + events
  * ownership: cross-user access returns 404 (not 403, no existence leak)
  * sessions CRUD + assign-to-project + invalid-cookie 422

Auth uses a real JWT minted with app.core.security.create_access_token (subject =
user id), so no password round-trip is needed. A throwaway user is seeded with a
COPY of an existing free-provider config (single app-wide Fernet key decrypts
it). All seeded data is deleted at the end.

Run (manages its own backend on port 8000):
    venv\\Scripts\\python.exe -m tests.e2e.run_http_api_e2e
"""

from __future__ import annotations

import asyncio
import http.server
import io
import logging
import os
import subprocess
import sys
import threading
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(REPO_ROOT / ".env", override=False)

API_BASE = "http://127.0.0.1:8000/api/v1"
FIXTURE_PORT = 9878
FIXTURE_BASE = f"http://127.0.0.1:{FIXTURE_PORT}"
BACKEND_STARTUP_TIMEOUT = 30

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("e2e.http")


# --------------------------------------------------------------------------- #
# fixture site
# --------------------------------------------------------------------------- #
def _items(rows: list[tuple[str, str, str]]) -> str:
    return "".join(
        f'<div class="item"><a class="name" href="{href}">{name}</a>'
        f'<span class="price">{price}</span></div>'
        for name, price, href in rows
    )


_NAV = '<nav><a href="/about">About</a><a href="/contact">Contact</a></nav>'
LISTING = (
    "<html><body>" + _NAV + '<div class="catalog">'
    + _items([("Widget Alpha", "$10.00", "/item/1"),
              ("Widget Beta", "$20.00", "/item/2"),
              ("Widget Gamma", "$30.00", "/item/3")])
    + '</div><a class="next" href="/?page=2">Next</a></body></html>'
)
PAGE2 = (
    "<html><body>" + _NAV + '<div class="catalog">'
    + _items([("Widget Delta", "$40.00", "/item/4"),
              ("Widget Epsilon", "$50.00", "/item/5")])
    + "</div></body></html>"
)
_DESC = ("This is a richly detailed product description that comfortably exceeds "
         "the minimum readable-content threshold so the CONTENT extractor returns "
         "a real record rather than treating the page as empty navigation. " * 2)
DETAIL = (
    "<html><body>" + _NAV
    + '<h1 class="title">Widget Alpha</h1><span class="price">$10.00</span>'
    + f'<div class="desc"><p>{_DESC}</p><p>Second paragraph of detail.</p></div>'
    + "</body></html>"
)
NAVPAGE = "<html><body>" + _NAV + "<p>About us.</p></body></html>"

ROUTES: dict[str, str] = {
    "/": LISTING,
    "/?page=2": PAGE2,
    "/item/1": DETAIL, "/item/2": DETAIL, "/item/3": DETAIL,
    "/item/4": DETAIL, "/item/5": DETAIL,
    "/about": NAVPAGE, "/contact": NAVPAGE,
}


class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, *a):  # noqa: A002, D102
        pass

    def do_GET(self) -> None:  # noqa: N802
        body = ROUTES.get(self.path) or ROUTES.get(self.path.split("?")[0])
        if self.path == "/robots.txt":
            self._send(200, "text/plain", b"User-agent: *\nAllow: /\n")
            return
        if body is None:
            self.send_response(404)
            self.end_headers()
            return
        self._send(200, "text/html; charset=utf-8", body.encode("utf-8"))

    def _send(self, code: int, ct: str, data: bytes) -> None:
        self.send_response(code)
        self.send_header("Content-Type", ct)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def start_fixture() -> http.server.HTTPServer:
    srv = http.server.HTTPServer(("127.0.0.1", FIXTURE_PORT), Handler)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    return srv


# --------------------------------------------------------------------------- #
# backend lifecycle (kills + owns port 8000 for the run)
# --------------------------------------------------------------------------- #
_proc: subprocess.Popen[bytes] | None = None


def _kill_8000() -> None:
    try:
        out = subprocess.run(["netstat", "-ano"], capture_output=True, text=True,
                             timeout=5).stdout
        for line in out.splitlines():
            if ":8000 " in line and "LISTENING" in line:
                pid = line.split()[-1]
                if pid.isdigit():
                    subprocess.run(["taskkill", "/PID", pid, "/F"],
                                   capture_output=True, timeout=5)
    except Exception:
        pass
    time.sleep(1.0)


def start_backend() -> bool:
    global _proc
    _kill_8000()
    env = os.environ.copy()
    env.update({
        "ALLOW_PRIVATE_NETWORK_URLS": "true",
        "ROBOTS_FAILURE_POLICY": "allow",
        "MIN_CRAWL_DELAY_MS": "0",
        "CRAWL_CONCURRENCY": "5",
        "ACCESS_TOKEN_EXPIRE_MINUTES": "120",
    })
    log_path = REPO_ROOT / "tests" / "e2e" / "backend_http_e2e.log"
    fh = open(log_path, "w")
    _proc = subprocess.Popen(
        [sys.executable, "-m", "uvicorn", "app.main:app",
         "--host", "127.0.0.1", "--port", "8000", "--no-access-log"],
        cwd=str(REPO_ROOT), env=env, stdout=fh, stderr=fh,
    )
    deadline = time.time() + BACKEND_STARTUP_TIMEOUT
    while time.time() < deadline:
        try:
            if httpx.get("http://127.0.0.1:8000/api/v1/health/live",
                         timeout=2.0).status_code == 200:
                return True
        except Exception:
            pass
        if _proc.poll() is not None:
            log.error("backend exited early - see %s", log_path)
            return False
        time.sleep(0.5)
    log.error("backend startup timeout - see %s", log_path)
    return False


def stop_backend() -> None:
    if _proc and _proc.poll() is None:
        _proc.terminate()
        try:
            _proc.wait(timeout=10)
        except subprocess.TimeoutExpired:
            _proc.kill()


def db_url() -> str:
    url = os.environ.get("DATABASE_URL", "")
    if not url:
        for line in (REPO_ROOT / ".env").read_text().splitlines():
            if line.startswith("DATABASE_URL="):
                url = line.split("=", 1)[1].strip()
    return url


# --------------------------------------------------------------------------- #
# seeding
# --------------------------------------------------------------------------- #
LIST_FIELDS = [
    {"name": "name", "label": "Name", "user_label": "Name", "selector": ".name",
     "type": "string", "selected": True},
    {"name": "price", "label": "Price", "user_label": "Price", "selector": ".price",
     "type": "string", "selected": True},
]
ANALYSIS = {"page_type": "product_listing", "confidence": 0.9,
            "repeated_item_selector": ".item", "candidate_fields": LIST_FIELDS,
            "detail_link_selector": ".name", "pagination_selector": ".next",
            "estimated_pages": 2, "warnings": []}


def _scope(mode: str, *, confirmed: bool = False) -> dict[str, Any]:
    return {
        "version": 1, "mode": mode,
        "status": "USER_CONFIRMED" if confirmed else "AI_SUGGESTED",
        "seed_url": None, "max_pages": 10, "max_depth": None,
        "include_patterns": [], "exclude_patterns": [],
        "pagination": {}, "link_rules": [], "ai_recommendation": None,
        "user_confirmed_at": datetime.now(timezone.utc).isoformat() if confirmed else None,
    }


@dataclass
class Seed:
    user_id: int
    other_user_id: int
    p_extract_current: int
    p_extract_pag: int
    p_noprev: int
    p_scope_gate: int
    p_zero: int
    p_failed: int
    p_delete: int
    p_completed_other: int


async def setup(url: str) -> Seed:
    from sqlalchemy import select
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.models.job import (ExtractionMode, ExtractionSpec, Project,
                                ProjectState, RenderMode, WorkflowMode)
    from app.models.provider_config import ProviderConfig
    from app.models.user import User

    eng = create_async_engine(url, echo=False, pool_size=2, max_overflow=0)
    Session = async_sessionmaker(eng, expire_on_commit=False)

    def mk_project(uid: int, path: str, state: ProjectState) -> Project:
        return Project(
            user_id=uid, url=f"{FIXTURE_BASE}{path}", normalized_url=f"{FIXTURE_BASE}{path}",
            state=state, extraction_mode=ExtractionMode.STRUCTURED,
            workflow_mode=WorkflowMode.GUIDED, render_mode=RenderMode.AUTO,
            confidence=0.9, warnings=[], analysis=dict(ANALYSIS),
            fetch_metadata={"status_code": 200, "content_type": "text/html"},
        )

    def mk_spec(pid: int, scope: dict, *, fields=None, page_limit=10) -> ExtractionSpec:
        return ExtractionSpec(
            project_id=pid, mode=ExtractionMode.STRUCTURED,
            fields=fields if fields is not None else [dict(f) for f in LIST_FIELDS],
            content_config={}, url_patterns=[], page_limit=page_limit,
            export_format="csv", crawl_scope=scope, quality_summary=None,
        )

    async with Session() as db:
        async with db.begin():
            u1 = User(email="e2e_http_main@example.com",
                      hashed_password="x", is_active=True)
            u2 = User(email="e2e_http_other@example.com",
                      hashed_password="x", is_active=True)
            db.add_all([u1, u2])
            await db.flush()

            # Copy a free provider so live analyze can run as u1. Prefer
            # gpt-oss (faster responses); fall back to any free model.
            free = (await db.execute(
                select(ProviderConfig).where(ProviderConfig.model.like("%gpt-oss%:free%"))
            )).scalars().first()
            if free is None:
                free = (await db.execute(
                    select(ProviderConfig).where(ProviderConfig.model.like("%:free%"))
                    .order_by(ProviderConfig.is_default.desc(), ProviderConfig.id)
                )).scalars().first()
            if free is not None:
                db.add(ProviderConfig(
                    user_id=u1.id, name="e2e-free", provider=free.provider,
                    model=free.model, api_key_encrypted=free.api_key_encrypted,
                    is_default=True, capability_flags=dict(free.capability_flags or {}),
                ))

            def add(p: Project, spec_scope: dict, **kw) -> int:
                db.add(p)
                return p

            p_cur = mk_project(u1.id, "/", ProjectState.ANALYSIS_READY)
            p_pag = mk_project(u1.id, "/", ProjectState.ANALYSIS_READY)
            p_np = mk_project(u1.id, "/", ProjectState.ANALYSIS_READY)
            p_sg = mk_project(u1.id, "/", ProjectState.ANALYSIS_READY)
            p_zero = mk_project(u1.id, "/", ProjectState.ANALYSIS_READY)
            p_fail = mk_project(u1.id, "/", ProjectState.FAILED)
            p_fail.error = "seeded failure"
            p_fail.error_code = "ALL_PAGES_FAILED"
            p_del = mk_project(u1.id, "/", ProjectState.COMPLETED)
            p_other = mk_project(u2.id, "/", ProjectState.COMPLETED)
            for p in (p_cur, p_pag, p_np, p_sg, p_zero, p_fail, p_del, p_other):
                db.add(p)
            await db.flush()

            db.add(mk_spec(p_cur.id, _scope("CURRENT_PAGE", confirmed=True)))
            db.add(mk_spec(p_pag.id, _scope("PAGINATION", confirmed=True)))
            db.add(mk_spec(p_np.id, _scope("CURRENT_PAGE", confirmed=True)))
            db.add(mk_spec(p_sg.id, _scope("FULL_SITE")))  # unconfirmed -> gate
            bogus = [{"name": "nope", "label": "Nope", "user_label": "Nope",
                      "selector": ".does-not-exist", "type": "string", "selected": True}]
            db.add(mk_spec(p_zero.id, _scope("CURRENT_PAGE", confirmed=True), fields=bogus))
            db.add(mk_spec(p_del.id, _scope("CURRENT_PAGE", confirmed=True)))
            db.add(mk_spec(p_other.id, _scope("CURRENT_PAGE", confirmed=True)))

            seed = Seed(
                user_id=u1.id, other_user_id=u2.id,
                p_extract_current=p_cur.id, p_extract_pag=p_pag.id,
                p_noprev=p_np.id, p_scope_gate=p_sg.id, p_zero=p_zero.id,
                p_failed=p_fail.id, p_delete=p_del.id, p_completed_other=p_other.id,
            )
    await eng.dispose()
    return seed


async def teardown(url: str, user_ids: list[int]) -> None:
    from sqlalchemy import text
    from sqlalchemy.ext.asyncio import create_async_engine

    eng = create_async_engine(url, echo=False, pool_size=2, max_overflow=0)
    ids = tuple(user_ids)
    proj_scoped = ["extracted_records", "extraction_runs", "frontier_previews",
                   "preview_results", "crawl_pages", "exports", "project_events",
                   "extraction_specs"]
    async with eng.begin() as conn:
        pid_sql = "SELECT id FROM projects WHERE user_id = ANY(:ids)"
        for tbl in proj_scoped:
            try:
                await conn.execute(
                    text(f"DELETE FROM {tbl} WHERE project_id IN ({pid_sql})"),
                    {"ids": list(ids)})
            except Exception as exc:  # noqa: BLE001
                log.warning("teardown %s: %s", tbl, exc)
        for tbl in ("browser_sessions", "provider_configs", "projects"):
            try:
                await conn.execute(
                    text(f"DELETE FROM {tbl} WHERE user_id = ANY(:ids)"),
                    {"ids": list(ids)})
            except Exception as exc:  # noqa: BLE001
                log.warning("teardown %s: %s", tbl, exc)
        await conn.execute(text("DELETE FROM users WHERE id = ANY(:ids)"),
                           {"ids": list(ids)})
    await eng.dispose()


# --------------------------------------------------------------------------- #
# client
# --------------------------------------------------------------------------- #
class Client:
    def __init__(self, token: str) -> None:
        self.h = {"Authorization": f"Bearer {token}"}
        self.c = httpx.Client(timeout=httpx.Timeout(60.0))

    def get(self, p, **kw):
        return self.c.get(f"{API_BASE}{p}", headers=self.h, **kw)

    def post(self, p, **kw):
        return self.c.post(f"{API_BASE}{p}", headers=self.h, **kw)

    def patch(self, p, **kw):
        return self.c.patch(f"{API_BASE}{p}", headers=self.h, **kw)

    def delete(self, p, **kw):
        return self.c.delete(f"{API_BASE}{p}", headers=self.h, **kw)


def code_of(r: httpx.Response) -> str:
    try:
        d = r.json().get("detail") or {}
        return d.get("error_code", "") if isinstance(d, dict) else ""
    except Exception:
        return ""


def poll_state(c: Client, pid: int, targets: set[str], timeout: float = 90.0) -> str:
    deadline = time.time() + timeout
    last = "?"
    while time.time() < deadline:
        r = c.get(f"/projects/{pid}")
        if r.status_code == 200:
            last = r.json().get("system_state", "?")
            if last in targets:
                return last
        time.sleep(1.0)
    return last


# --------------------------------------------------------------------------- #
# results
# --------------------------------------------------------------------------- #
@dataclass
class Result:
    name: str
    status: str = "PASS"          # PASS | FAIL | LIMIT | ENV
    evidence: list[str] = field(default_factory=list)
    failures: list[str] = field(default_factory=list)

    def ok(self, m: str) -> None:
        self.evidence.append(m)

    def bad(self, m: str) -> None:
        self.failures.append(m)
        self.status = "FAIL"


# --------------------------------------------------------------------------- #
# scenarios
# --------------------------------------------------------------------------- #
def sc_live_extract_current(c: Client, sd: Seed) -> Result:
    r = Result("LIVE extract CURRENT_PAGE -> COMPLETED -> records -> export")
    pid = sd.p_extract_current
    try:
        g = c.get(f"/projects/{pid}")
        if g.status_code != 200:
            r.bad(f"GET project {g.status_code}")
            return r
        r.ok("GET project 200")

        pv = c.post(f"/projects/{pid}/preview")
        if pv.status_code != 200:
            r.bad(f"preview {pv.status_code}: {pv.text[:160]}")
            return r
        n = len(pv.json().get("sample_records") or [])
        if n < 1:
            r.bad(f"preview produced {n} sample records")
            return r
        r.ok(f"preview 200 with {n} sample records")

        fp = c.post(f"/projects/{pid}/frontier-preview")
        if fp.status_code != 201:
            r.bad(f"frontier-preview {fp.status_code}: {fp.text[:160]}")
        else:
            r.ok(f"frontier-preview 201 ({len(fp.json().get('included_urls', []))} included)")

        ex = c.post(f"/projects/{pid}/extract", json={})
        if ex.status_code not in (200, 202):
            r.bad(f"extract {ex.status_code} code={code_of(ex)}: {ex.text[:160]}")
            return r
        r.ok(f"extract accepted ({ex.status_code})")

        state = poll_state(c, pid, {"COMPLETED", "FAILED"})
        if state != "COMPLETED":
            r.bad(f"extract did not COMPLETE (state={state})")
            return r
        r.ok("project reached COMPLETED")

        rp = c.get(f"/projects/{pid}/records-page?skip=0&limit=100")
        total = rp.json().get("total")
        if total != 3:
            r.bad(f"records total={total} (want 3)")
        else:
            r.ok("records-page total=3")

        for fmt, check in (("csv", None), ("json", None), ("xlsx", None)):
            ex_r = c.get(f"/projects/{pid}/export?format={fmt}")
            if ex_r.status_code != 200:
                r.bad(f"export {fmt} {ex_r.status_code}")
                continue
            if fmt == "csv":
                lines = [l for l in ex_r.text.splitlines() if l]
                ok = len(lines) == 4 and "Widget Alpha" in ex_r.text
                r.ok(f"export csv 4 lines") if ok else r.bad(f"csv bad: {lines[:2]}")
            elif fmt == "json":
                data = ex_r.json()
                r.ok("export json 3 rows") if len(data) == 3 else r.bad(f"json {len(data)}")
            else:
                wb = load_workbook(io.BytesIO(ex_r.content))
                rows = wb.active.max_row
                r.ok(f"export xlsx {rows} rows") if rows == 4 else r.bad(f"xlsx {rows}")
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_live_extract_pagination(c: Client, sd: Seed) -> Result:
    r = Result("LIVE extract PAGINATION (multi-page) -> COMPLETED")
    pid = sd.p_extract_pag
    try:
        pv = c.post(f"/projects/{pid}/preview")
        if pv.status_code != 200:
            r.bad(f"preview {pv.status_code}: {pv.text[:160]}")
            return r
        r.ok(f"preview 200 ({len(pv.json().get('sample_records') or [])} samples)")
        ex = c.post(f"/projects/{pid}/extract", json={})
        if ex.status_code not in (200, 202):
            r.bad(f"extract {ex.status_code} code={code_of(ex)}")
            return r
        state = poll_state(c, pid, {"COMPLETED", "FAILED"})
        if state != "COMPLETED":
            r.bad(f"state={state}")
            return r
        total = c.get(f"/projects/{pid}/records-page?skip=0&limit=100").json().get("total")
        if total and total > 3:
            r.ok(f"PAGINATION followed >1 page (records={total})")
        else:
            r.bad(f"PAGINATION records={total} (want >3 across 2 pages)")
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_gates(c: Client, sd: Seed) -> Result:
    r = Result("EXTRACT GATES: NO_PREVIEW / STALE_PREVIEW / ZERO_PREVIEW_RECORDS / SCOPE_NOT_CONFIRMED")
    try:
        # NO_PREVIEW
        g1 = c.post(f"/projects/{sd.p_noprev}/extract", json={})
        (r.ok("NO_PREVIEW 409") if g1.status_code == 409 and code_of(g1) == "NO_PREVIEW"
         else r.bad(f"NO_PREVIEW: {g1.status_code}/{code_of(g1)}"))

        # SCOPE_NOT_CONFIRMED (FULL_SITE unconfirmed, bypass soft gates)
        g2 = c.post(f"/projects/{sd.p_scope_gate}/extract", json={"extract_anyway": True})
        (r.ok("SCOPE_NOT_CONFIRMED 409")
         if g2.status_code == 409 and code_of(g2) == "SCOPE_NOT_CONFIRMED"
         else r.bad(f"SCOPE_NOT_CONFIRMED: {g2.status_code}/{code_of(g2)}"))

        # ZERO_PREVIEW_RECORDS (bogus selector -> preview 0 -> extract hard-gated)
        pvz = c.post(f"/projects/{sd.p_zero}/preview")
        if pvz.status_code == 200 and len(pvz.json().get("sample_records") or []) == 0:
            gz = c.post(f"/projects/{sd.p_zero}/extract", json={})
            (r.ok("ZERO_PREVIEW_RECORDS 409")
             if gz.status_code == 409 and code_of(gz) == "ZERO_PREVIEW_RECORDS"
             else r.bad(f"ZERO_PREVIEW_RECORDS: {gz.status_code}/{code_of(gz)}"))
        else:
            r.bad(f"zero-preview setup: preview {pvz.status_code}, "
                  f"{len(pvz.json().get('sample_records') or [])} samples")

        # STALE_PREVIEW: preview, then change fields, then extract
        pid = sd.p_noprev
        c.post(f"/projects/{pid}/preview")
        patched = c.patch(f"/projects/{pid}/spec", json={"fields": [
            {"name": "name", "label": "Renamed", "user_label": "Renamed",
             "selector": ".name", "type": "string", "selected": True}]})
        if patched.status_code == 200:
            gs = c.post(f"/projects/{pid}/extract", json={})
            (r.ok("STALE_PREVIEW 409")
             if gs.status_code == 409 and code_of(gs) == "STALE_PREVIEW"
             else r.bad(f"STALE_PREVIEW: {gs.status_code}/{code_of(gs)}"))
        else:
            r.bad(f"stale setup patch {patched.status_code}")
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_ownership(c_other: Client, sd: Seed) -> Result:
    r = Result("OWNERSHIP: cross-user access -> 404 (no existence leak)")
    pid = sd.p_extract_current  # owned by user1; c_other is user2
    try:
        for label, resp in [
            ("GET project", c_other.get(f"/projects/{pid}")),
            ("records-page", c_other.get(f"/projects/{pid}/records-page?skip=0&limit=10")),
            ("extract", c_other.post(f"/projects/{pid}/extract", json={})),
            ("events", c_other.get(f"/projects/{pid}/events")),
            ("delete", c_other.delete(f"/projects/{pid}")),
        ]:
            (r.ok(f"{label} -> 404") if resp.status_code == 404
             else r.bad(f"{label} -> {resp.status_code} (want 404)"))
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_sessions(c: Client, sd: Seed) -> Result:
    r = Result("SESSIONS: create / list / assign-to-project / invalid-cookie 422 / delete")
    try:
        cr = c.post("/sessions", json={
            "name": "e2e-sess", "domain": "127.0.0.1",
            "cookies_raw": "sid=abc123; theme=dark"})
        if cr.status_code != 201:
            r.bad(f"create session {cr.status_code}: {cr.text[:160]}")
            return r
        sid = cr.json()["id"]
        r.ok(f"create session 201 (id={sid})")

        lst = c.get("/sessions")
        (r.ok(f"list sessions ({len(lst.json())})") if lst.status_code == 200
         else r.bad(f"list {lst.status_code}"))

        asg = c.patch(f"/projects/{sd.p_delete}/session?browser_session_id={sid}")
        (r.ok("assign session to project 200") if asg.status_code == 200
         else r.bad(f"assign {asg.status_code}: {asg.text[:160]}"))

        clr = c.patch(f"/projects/{sd.p_delete}/session")
        (r.ok("clear session 200") if clr.status_code == 200
         else r.bad(f"clear {clr.status_code}"))

        bad = c.post("/sessions", json={
            "name": "bad", "domain": "127.0.0.1", "cookies_raw": '[{"no":"name"}]'})
        (r.ok("invalid cookie -> 422") if bad.status_code == 422
         else r.bad(f"invalid cookie -> {bad.status_code} (want 422)"))

        dl = c.delete(f"/sessions/{sid}")
        (r.ok("delete session 204") if dl.status_code == 204
         else r.bad(f"delete session {dl.status_code}"))
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_lifecycle(c: Client, sd: Seed) -> Result:
    r = Result("LIFECYCLE: retry(FAILED) / cancel(non-active 409) / events / delete(204)")
    try:
        # retry a FAILED project (provider seeded -> should be accepted)
        rt = c.post(f"/projects/{sd.p_failed}/retry", json={})
        if rt.status_code == 200:
            r.ok("retry FAILED project -> 200")
        elif rt.status_code == 409:
            r.status = "LIMIT"
            r.evidence.append(f"retry -> 409 {code_of(rt)} (provider/state dependent)")
        else:
            r.bad(f"retry -> {rt.status_code}: {rt.text[:160]}")

        # cancel a non-active (COMPLETED) project -> 409
        cn = c.post(f"/projects/{sd.p_completed_other}/cancel")  # not owned -> 404 actually
        # use an owned completed project instead:
        cn = c.post(f"/projects/{sd.p_delete}/cancel")
        (r.ok("cancel non-active -> 409") if cn.status_code == 409
         else r.bad(f"cancel non-active -> {cn.status_code} (want 409)"))

        ev = c.get(f"/projects/{sd.p_delete}/events")
        (r.ok(f"events -> 200 ({len(ev.json())})") if ev.status_code == 200
         else r.bad(f"events -> {ev.status_code}"))

        dl = c.delete(f"/projects/{sd.p_delete}")
        (r.ok("delete completed project -> 204") if dl.status_code == 204
         else r.bad(f"delete -> {dl.status_code}: {dl.text[:160]}"))
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_live_analyze(c: Client) -> Result:
    """Live analyze through the user's free provider. Provider quality/availability
    is classified honestly as ENV/LIMIT, not a pipeline FAIL."""
    r = Result("LIVE analyze (free provider) -> ANALYSIS_READY")
    try:
        an = c.post("/projects/analyze", json={"url": f"{FIXTURE_BASE}/"})
        if an.status_code == 409 and code_of(an) == "NO_PROVIDER_CONFIGURED":
            r.status = "ENV"
            r.evidence.append("no provider configured for test user (analyze 409)")
            return r
        if an.status_code != 202:
            r.bad(f"analyze -> {an.status_code}: {an.text[:160]}")
            return r
        pid = an.json()["id"]
        r.ok(f"analyze accepted 202 (project {pid})")
        state = poll_state(c, pid, {"ANALYSIS_READY", "AWAITING_SETUP", "FAILED"}, 120.0)
        if state in {"ANALYSIS_READY", "AWAITING_SETUP"}:
            proj = c.get(f"/projects/{pid}").json()
            nfields = len((proj.get("analysis") or {}).get("candidate_fields") or [])
            r.ok(f"reached {state} with {nfields} candidate field(s)")
        else:
            r.status = "ENV"
            err = c.get(f"/projects/{pid}").json().get("error_code")
            r.evidence.append(f"analyze did not finish (state={state}, err={err}) "
                              f"- free-model availability")
        c.delete(f"/projects/{pid}")
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}\n{traceback.format_exc()}")
    return r


def sc_auth() -> Result:
    r = Result("AUTH: missing/invalid token -> 401")
    try:
        a = httpx.get(f"{API_BASE}/projects", timeout=10.0)
        (r.ok("no token -> 401") if a.status_code == 401
         else r.bad(f"no token -> {a.status_code}"))
        b = httpx.get(f"{API_BASE}/projects",
                      headers={"Authorization": "Bearer not.a.jwt"}, timeout=10.0)
        (r.ok("garbage token -> 401") if b.status_code == 401
         else r.bad(f"garbage token -> {b.status_code}"))
    except Exception as exc:  # noqa: BLE001
        r.bad(f"exception: {exc}")
    return r


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #
def main() -> int:
    from app.core.security import create_access_token

    fixture = start_fixture()
    time.sleep(0.2)
    if httpx.get(f"{FIXTURE_BASE}/", timeout=3.0).status_code != 200:
        log.error("fixture server not serving")
        return 1
    log.info("fixture up at %s", FIXTURE_BASE)

    url = db_url()
    log.info("seeding...")
    sd = asyncio.run(setup(url))
    log.info("seeded user_id=%s other=%s", sd.user_id, sd.other_user_id)

    results: list[Result] = []
    try:
        if not start_backend():
            return 1
        token = create_access_token(subject=sd.user_id)
        token_other = create_access_token(subject=sd.other_user_id)
        c = Client(token)
        c_other = Client(token_other)

        scenarios = [
            lambda: sc_auth(),
            lambda: sc_live_extract_current(c, sd),
            lambda: sc_live_extract_pagination(c, sd),
            lambda: sc_gates(c, sd),
            lambda: sc_ownership(c_other, sd),
            lambda: sc_sessions(c, sd),
            lambda: sc_lifecycle(c, sd),
            lambda: sc_live_analyze(c),
        ]
        for fn in scenarios:
            res = fn()
            results.append(res)
            log.info("--- %s : %s ---", res.name, res.status)
            for e in res.evidence:
                log.info("   OK  %s", e)
            for fa in res.failures:
                log.warning("   BAD %s", fa)
    finally:
        stop_backend()
        try:
            asyncio.run(teardown(url, [sd.user_id, sd.other_user_id]))
            log.info("teardown complete")
        except Exception as exc:  # noqa: BLE001
            log.error("teardown failed: %s", exc)
        fixture.shutdown()

    print("\n" + "=" * 78)
    print("LAYER B - HTTP API E2E RESULTS")
    print("=" * 78)
    counts = {"PASS": 0, "FAIL": 0, "LIMIT": 0, "ENV": 0}
    for res in results:
        counts[res.status] = counts.get(res.status, 0) + 1
        print(f"[{res.status:5}] {res.name}")
        for fa in res.failures:
            print(f"          - {fa.splitlines()[0]}")
    print("-" * 78)
    print(f"PASS={counts['PASS']} FAIL={counts['FAIL']} "
          f"LIMIT={counts['LIMIT']} ENV={counts['ENV']}")
    print("=" * 78)
    return counts["FAIL"]


if __name__ == "__main__":
    raise SystemExit(main())
